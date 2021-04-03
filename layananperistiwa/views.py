import pytz
from csv import DictWriter
from io import BytesIO, StringIO
from datetime import datetime, date
from rest_framework import permissions
from rest_framework import filters
from openpyxl import Workbook

from django.conf import settings
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.exceptions import NotFound, APIException
from dynamic_rest.viewsets import DynamicModelViewSet

from api_sad_sig.util import CustomView
from users.permissions import IsAdminUserOrReadOnly
from v1.models import SadPenduduk

from .utils import render_mail
from .models import (
    SuratDomisili,
    SuratKelahiran,
    SuratSkck,
    SadKelahiran,
    SadKematian,
    SadLahirmati,
    JenisPindah,
    AlasanPindah,
    KlasifikasiPindah,
    StatusKKPindah,
    StatusKKTinggal,
    SadPindahKeluar,
    SadPindahMasuk,
    SadPecahKK,
)
from .serializers import (
    AdminSuratDomisiliSerializer,
    AdminSuratKelahiranSerializer,
    AdminSuratSkckSerializer,
    SuratDomisiliSerializer,
    SuratKelahiranSerializer,
    SuratSkckSerializer,
    SadKelahiranSerializer,
    SadKematianSerializer,
    SadLahirmatiSerializer,
    AlasanPindahSerializer,
    KlasifikasiPindahSerializer,
    JenisPindahSerializer,
    StatusKKPindahSerializer,
    StatusKKTinggalSerializer,
    SadPindahKeluarSerializer,
    SadPindahMasukSerializer,
    SadPecahKKSerializer,
    LaporanKelahiranSerializer,
    LaporanKematianSerializer,
    LaporanMonografiSerializer,
)


class SuratKelahiranViewSet(DynamicModelViewSet):
    queryset = SuratKelahiran.objects.all().order_by("-id")
    permission_classes = [permissions.IsAuthenticatedOrReadOnly]

    filter_backends = [filters.SearchFilter]
    search_fields = ["nama"]

    def get_serializer_class(self):
        if self.request.user.groups.first().name == "admin":
            return AdminSuratKelahiranSerializer
        return SuratKelahiranSerializer

    @action(detail=True, methods=["get"])
    def print(self, request, pk=None):
        data = self.get_object()
        pdf = render_mail("skl", data)
        return HttpResponse(pdf, content_type="application/pdf")


class SuratSkckViewSet(DynamicModelViewSet):
    queryset = SuratSkck.objects.all().order_by("-id")
    permission_classes = [permissions.IsAuthenticatedOrReadOnly]

    filter_backends = [filters.SearchFilter]
    search_fields = ["nama"]

    def get_serializer_class(self):
        if self.request.user.groups.first().name == "admin":
            return AdminSuratSkckSerializer
        return SuratSkckSerializer

    @action(detail=True, methods=["get"])
    def print(self, request, pk=None):
        data = self.get_object()
        pdf = render_mail("skck", data)
        return HttpResponse(pdf, content_type="application/pdf")


class SuratDomisiliViewSet(DynamicModelViewSet):
    queryset = SuratDomisili.objects.all().order_by("-id")
    permission_classes = [permissions.IsAuthenticatedOrReadOnly]

    filter_backends = [filters.SearchFilter]
    search_fields = ["nama"]

    def get_serializer_class(self):
        if self.request.user.groups.first().name == "admin":
            return AdminSuratDomisiliSerializer
        return SuratDomisiliSerializer

    @action(detail=True, methods=["get"])
    def print(self, request, pk=None):
        data = self.get_object()
        pdf = render_mail("skd", data)
        return HttpResponse(pdf, content_type="application/pdf")


def string_to_date(text):
    return datetime.strptime(text, "%Y-%m-%d").astimezone(
        pytz.timezone(settings.TIME_ZONE)
    )


class LaporanMonografiViewSet(DynamicModelViewSet):
    queryset = SadPenduduk.objects.order_by("id").all()
    serializer_class = LaporanMonografiSerializer
    permission_classes = [IsAdminUserOrReadOnly]

    def get_serializer_class(self):
        print(self.action)
        if self.action not in ["list", "create", "print"]:
            raise NotFound("Operasi ini tidak tersedia")
        return self.serializer_class

    @action(detail=False, methods=["get"])
    def print(self, request):
        data = self.get_queryset()
        pdf = render_mail("monografi", data)
        return HttpResponse(pdf, content_type="application/pdf")

    @action(detail=False, methods=["get"])
    def csv(self, request):
        extras = {
            "Nomor NIK": "nik",
            "Nama": "nama",
            "Jenis Kelamin": "jk",
            "Tanggal Lahir": "tgl_lahir",
            "Pendidikan": "pendidikan",
            "Pekerjaan": "pekerjaan",
        }
        data = (
            self.get_queryset()
            .extra(select=extras)
            .values(*extras.keys())
            .all()
        )

        output = StringIO()
        writer = DictWriter(output, list(extras.keys()))
        writer.writeheader()
        writer.writerows(data)

        response = HttpResponse(output.getvalue(), content_type="text/csv")
        response[
            "Content-Disposition"
        ] = 'attachment; filename="Monografi.csv"'
        return response

    @action(detail=False, methods=["get"])
    def excel(self, request):
        extras = {
            "Nomor NIK": "nik",
            "Nama": "nama",
            "Jenis Kelamin": "jk",
            "Tanggal Lahir": "tgl_lahir",
            "Pendidikan": "pendidikan",
            "Pekerjaan": "pekerjaan",
        }
        data = (
            self.get_queryset()
            .extra(select=extras)
            .values(*extras.keys())
            .all()
        )

        workbook = Workbook()
        sheet = workbook.active

        headers = [i for i in extras.keys()]
        for index, value in enumerate(headers):
            sheet.cell(row=1, column=index + 1).value = value

        for i, x in enumerate(data):
            for idx, value in enumerate(x.values()):
                sheet.cell(row=i + 2, column=idx + 1).value = value

        output = BytesIO()
        workbook.save(output)
        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.ms-excel",
        )
        response[
            "Content-Disposition"
        ] = 'attachment; filename="Monografi.xls"'
        return response

    def get_queryset(self):
        start = self.request.query_params.get("start")
        end = self.request.query_params.get("end")
        if not start or not end:
            raise APIException(
                "Need start date and end date for filtering", 400
            )
        start_date = string_to_date(start)
        end_date = string_to_date(end)

        queryset = SadPenduduk.objects.filter(
            created_at__gte=start_date, created_at__lte=end_date
        )

        allowed_param = ("jk", "pekerjaan", "status_kawin")
        param = self.request.query_params.get("param")
        value = self.request.query_params.get("value")
        if param not in ("dpt", "child") and not value:
            raise APIException("Param needs Value", 400)

        if param and value and param in allowed_param:
            dict_params = {param: value}
            queryset = queryset.filter(**dict_params)

        if param == "age" and not value.isdigit():
            raise APIException("Value needs to be integer", 400)
        if param == "age":
            today = date.today()
            min_date = date(
                today.year - (int(value) + 1), today.month, today.day
            )
            max_date = date(today.year - int(value), today.month, today.day)
            queryset = queryset.filter(
                tgl_lahir__gt=min_date, tgl_lahir__lte=max_date
            )
        if param == "dpt":
            today = date.today()
            max_date = date(today.year - 17, today.month, today.day)
            queryset = queryset.filter(tgl_lahir__lte=max_date)
        if param == "child":
            today = date.today()
            min_date = date(today.year - 12, today.month, today.day)
            max_date = date(today.year - 5, today.month, today.day)
            queryset = queryset.filter(
                tgl_lahir__gt=min_date, tgl_lahir__lte=max_date
            )

        return queryset.order_by("id")


class LaporanKelahiranViewSet(CustomView):
    serializer_class = LaporanKelahiranSerializer
    permission_classes = [IsAdminUserOrReadOnly]

    def get_queryset(self):
        quarter = self.request.query_params.get("triwulan", None)
        year, month = quarter.split("-")
        if not year or not month:
            raise APIException('Wrong format for "triwulan"', 400)
        if not year.isdigit() or not month.isdigit():
            raise APIException("Year and Month need to be integer format", 400)

        start = 3 * (int(month) - 1) + 1
        end = start + 3

        return SadKelahiran.objects.filter(created_at__year=int(year)).filter(
            created_at__month__in=tuple(i for i in range(start, end))
        )

    @action(detail=False, methods=["get"])
    def print(self, request):
        data = self.get_queryset()
        pdf = render_mail("triwulan_kelahiran", data)
        return HttpResponse(pdf, content_type="application/pdf")

    @action(detail=False, methods=["get"])
    def csv(self, request):
        extras = {
            "Nama": "nama",
            "Jenis Kelamin": "jenis_kelamin",
            "Tanggal Kelahiran": "tanggal_kelahiran",
            "Tempat Kelahiran": "tempat_kelahiran",
            "NIK Ayah": "nik_ayah",
            "Nama Ayah": "nama_ayah",
            "NIK Ibu": "nik_ibu",
            "Nama Ibu": "nama_ibu",
        }
        data = (
            self.get_queryset()
            .extra(select=extras)
            .values(*extras.keys())
            .all()
        )

        output = StringIO()
        writer = DictWriter(output, list(extras.keys()))
        writer.writeheader()
        writer.writerows(data)

        response = HttpResponse(output.getvalue(), content_type="text/csv")
        response[
            "Content-Disposition"
        ] = 'attachment; filename="Triwulan Kelahiran.csv"'
        return response

    @action(detail=False, methods=["get"])
    def excel(self, request):
        extras = {
            "Nama": "nama",
            "Jenis Kelamin": "jenis_kelamin",
            "Tanggal Kelahiran": "tanggal_kelahiran",
            "Tempat Kelahiran": "tempat_kelahiran",
            "NIK Ayah": "nik_ayah",
            "Nama Ayah": "nama_ayah",
            "NIK Ibu": "nik_ibu",
            "Nama Ibu": "nama_ibu",
        }
        data = (
            self.get_queryset()
            .extra(select=extras)
            .values(*extras.keys())
            .all()
        )

        workbook = Workbook()
        sheet = workbook.active

        headers = [i for i in extras.keys()]
        for index, value in enumerate(headers):
            sheet.cell(row=1, column=index + 1).value = value

        for i, x in enumerate(data):
            for idx, value in enumerate(x.values()):
                sheet.cell(row=i + 2, column=idx + 1).value = value

        output = BytesIO()
        workbook.save(output)
        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.ms-excel",
        )
        response[
            "Content-Disposition"
        ] = 'attachment; filename="TriwulanKelahiran.xls"'
        return response


class SadKelahiranViewSet(CustomView):
    queryset = SadKelahiran.objects.all().order_by("-id")
    serializer_class = SadKelahiranSerializer
    permission_classes = [IsAdminUserOrReadOnly]

    filter_backends = [filters.SearchFilter]
    search_fields = ["nama", "nama_ayah", "nama_ibu"]

    @action(detail=False, methods=["get"])
    def ekspor(self, request):
        extras = {
            "Nama": "nama",
            "Jenis Kelamin": "jenis_kelamin",
            "Tempat Kelahiran": "tempat_kelahiran",
            "Tanggal Kelahiran": "tanggal_kelahiran",
            "Jam Kelahiran": "jam",
            "Jenis Kelahiran": "jenis_kelahiran",
            "Tempat Dilahirkan": "tempat_dilahirkan",
            "Penolong Kelahiran": "penolong_kelahiran",
            "Berat Bayi": "berat_bayi",
            "Panjang Bayi": "panjang_bayi",
            "NIK Ayah": "nik_ayah",
            "Nama Ayah": "nama_ayah",
            "NIK Ibu": "nik_ibu",
            "Nama Ibu": "nama_ibu",
            "Anak Ke": "kelahiran_ke",
        }
        data = (
            self.get_queryset()
            .extra(select=extras)
            .values(*extras.keys())
            .all()
        )

        workbook = Workbook()
        sheet = workbook.active

        headers = [i for i in extras.keys()]
        for index, value in enumerate(headers):
            sheet.cell(row=1, column=index + 1).value = value

        for i, x in enumerate(data):
            for idx, value in enumerate(x.values()):
                sheet.cell(row=i + 2, column=idx + 1).value = value

        output = BytesIO()
        workbook.save(output)
        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.ms-excel",
        )
        response[
            "Content-Disposition"
        ] = 'attachment; filename="DataKelahiran.xlsx"'
        return response


class LaporanKematianViewSet(DynamicModelViewSet):
    queryset = SadKematian.objects.all().order_by("id")
    serializer_class = LaporanKematianSerializer
    permission_classes = [IsAdminUserOrReadOnly]

    def get_serializer_class(self):
        if self.action not in ["list", "create", "print"]:
            raise NotFound("Operasi ini tidak tersedia")
        return self.serializer_class

    def get_queryset(self):
        quarter = self.request.query_params.get("triwulan", "print")
        if not quarter:
            raise APIException("Need triwulan parameter")

        year, month = quarter.split("-")
        if not year or not month:
            raise APIException('Wrong format for "triwulan"', 400)
        if not year.isdigit() or not month.isdigit():
            raise APIException("Year and Month need to be integer format", 400)

        start = 3 * (int(month) - 1) + 1
        end = start + 3

        return (
            SadKematian.objects.filter(created_at__year=int(year))
            .filter(created_at__month__in=tuple(i for i in range(start, end)))
            .order_by("id")
        )

    @action(detail=False, methods=["get"])
    def print(self, request):
        data = self.get_queryset()
        pdf = render_mail("triwulan_kematian", data)
        return HttpResponse(pdf, content_type="application/pdf")

    @action(detail=False, methods=["get"])
    def csv(self, request):
        extras = {
            "NIK": "sad_penduduk.nik",
            "Nama": "sad_penduduk.nama",
            "Tanggal Kematian": "tanggal_kematian",
            "Tempat Kematian": "tempat_kematian",
            "Sebab Kematian": "sebab_kematian",
            "Yang Menerangkan": "yang_menerangkan",
            "Nama Pelapor": "nama_pelapor",
        }
        data = (
            self.get_queryset()
            .extra(select=extras, tables=("sad_penduduk",))
            .values(*extras.keys())
            .all()
        )

        output = StringIO()
        writer = DictWriter(output, list(extras.keys()))
        writer.writeheader()
        writer.writerows(data)

        response = HttpResponse(output.getvalue(), content_type="text/csv")
        response[
            "Content-Disposition"
        ] = 'attachment; filename="Triwulan Kematian.csv"'
        return response

    @action(detail=False, methods=["get"])
    def excel(self, request):
        extras = {
            "NIK": "sad_penduduk.nik",
            "Nama": "sad_penduduk.nama",
            "Tanggal Kematian": "tanggal_kematian",
            "Tempat Kematian": "tempat_kematian",
            "Sebab Kematian": "sebab_kematian",
            "Yang Menerangkan": "yang_menerangkan",
            "Nama Pelapor": "nama_pelapor",
        }
        data = (
            self.get_queryset()
            .extra(select=extras, tables=("sad_penduduk",))
            .values(*extras.keys())
            .all()
        )

        workbook = Workbook()
        sheet = workbook.active

        headers = [i for i in extras.keys()]
        for index, value in enumerate(headers):
            sheet.cell(row=1, column=index + 1).value = value

        for i, x in enumerate(data):
            for idx, value in enumerate(x.values()):
                sheet.cell(row=i + 2, column=idx + 1).value = value

        output = BytesIO()
        workbook.save(output)
        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.ms-excel",
        )
        response[
            "Content-Disposition"
        ] = 'attachment; filename="TriwulanKematian.xls"'
        return response


class SadKematianViewSet(CustomView):
    queryset = SadKematian.objects.all().order_by("id")
    serializer_class = SadKematianSerializer
    permission_classes = [IsAdminUserOrReadOnly]

    filter_backends = [filters.SearchFilter]
    search_fields = ["nama", "sebab_kematian"]

    @action(detail=False, methods=["get"])
    def ekspor(self, request):
        extras = {
            "NIK": "sad_penduduk.nik",
            "Nama": "sad_penduduk.nama",
            "Jenis Kelamin": "sad_penduduk.jk",
            "Tempat Lahir": "sad_penduduk.tempat_lahir",
            "Tanggal Lahir": "sad_penduduk.tgl_lahir",
            "Tempat Kematian": "tempat_kematian",
            "Tanggal Kematian": "tanggal_kematian",
            "Jam Kematian": "jam",
            "Sebab Kematian": "sebab_kematian",
            "Yang Menerangkan": "yang_menerangkan",
            "Nama Pelapor": "nama_pelapor",
        }
        data = (
            self.get_queryset()
            .extra(select=extras, tables=("sad_penduduk",))
            .values(*extras.keys())
            .all()
        )

        workbook = Workbook()
        sheet = workbook.active

        headers = [i for i in extras.keys()]
        for index, value in enumerate(headers):
            sheet.cell(row=1, column=index + 1).value = value

        for i, x in enumerate(data):
            for idx, value in enumerate(x.values()):
                sheet.cell(row=i + 2, column=idx + 1).value = value

        output = BytesIO()
        workbook.save(output)
        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.ms-excel",
        )
        response[
            "Content-Disposition"
        ] = 'attachment; filename="DataKematian.xlsx"'
        return response


class SadLahirmatiViewSet(CustomView):
    queryset = SadLahirmati.objects.all().order_by("id")
    serializer_class = SadLahirmatiSerializer
    permission_classes = [IsAdminUserOrReadOnly]

    filter_backends = [filters.SearchFilter]
    search_fields = ["nama"]

    @action(detail=False, methods=["get"])
    def ekspor(self, request):
        extras = {
            "Lama Kandungan": "lama_kandungan",
            "Jenis Kelamin": "jenis_kelamin",
            "Tanggal Lahir": "tanggal_lahir",
            "Tempat Kelahiran": "tempat_kelahiran",
            "Jenis Kelahiran": "jenis_kelahiran",
            "Tempat Dilahirkan": "tempat_dilahirkan",
            "Penolong Kelahiran": "penolong_kelahiran",
            "Sebab Lahir Mati": "sebab_lahirmati",
            "NIK Ayah": "nik_ayah",
            "Nama Ayah": "nama_ayah",
            "NIK Ibu": "nik_ibu",
            "Nama Ibu": "nama_ibu",
            "Anak Ke": "kelahiran_ke",
            "Nama Pelapor": "nama_pelapor",
        }
        data = (
            self.get_queryset()
            .extra(select=extras)
            .values(*extras.keys())
            .all()
        )

        workbook = Workbook()
        sheet = workbook.active

        headers = [i for i in extras.keys()]
        for index, value in enumerate(headers):
            sheet.cell(row=1, column=index + 1).value = value

        for i, x in enumerate(data):
            for idx, value in enumerate(x.values()):
                sheet.cell(row=i + 2, column=idx + 1).value = value

        output = BytesIO()
        workbook.save(output)
        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.ms-excel",
        )
        response[
            "Content-Disposition"
        ] = 'attachment; filename="DataLahirMati.xlsx"'
        return response


class JenisPindahViewSet(CustomView):
    queryset = JenisPindah.objects.all()
    serializer_class = JenisPindahSerializer
    permission_classes = [IsAdminUserOrReadOnly]


class AlasanPindahViewSet(CustomView):
    queryset = AlasanPindah.objects.all()
    serializer_class = AlasanPindahSerializer
    permission_classes = [IsAdminUserOrReadOnly]


class KlasifikasiPindahViewSet(CustomView):
    queryset = KlasifikasiPindah.objects.all()
    serializer_class = KlasifikasiPindahSerializer
    permission_classes = [IsAdminUserOrReadOnly]


class StatusKKTinggalViewSet(CustomView):
    queryset = StatusKKTinggal.objects.all()
    serializer_class = StatusKKTinggalSerializer
    permission_classes = [IsAdminUserOrReadOnly]


class StatusKKPindahViewSet(CustomView):
    queryset = StatusKKPindah.objects.all()
    serializer_class = StatusKKPindahSerializer
    permission_classes = [IsAdminUserOrReadOnly]


class SadPindahKeluarViewSet(CustomView):
    queryset = SadPindahKeluar.objects.all().order_by("id")
    serializer_class = SadPindahKeluarSerializer
    permission_classes = [IsAdminUserOrReadOnly]

    def retrieve(self, request, pk=None):
        queryset = SadPindahKeluar.objects.all()
        sad_pindah = get_object_or_404(queryset, pk=pk)
        serializer = SadPindahKeluarSerializer(sad_pindah)
        data = serializer.data

        penduduk_s = sad_pindah.anggota_keluar()
        penduduk_data = []
        for item in penduduk_s:
            temp_data = {"nik": item.nik, "nama": item.nama}
            penduduk_data.append(temp_data)

        data["anggota_keluar"] = penduduk_data
        return Response(data)


class SadPindahMasukViewSet(CustomView):
    queryset = SadPindahMasuk.objects.all().order_by("id")
    serializer_class = SadPindahMasukSerializer
    permission_classes = [IsAdminUserOrReadOnly]

    filter_backends = [filters.SearchFilter]
    search_fields = ["no_kk"]


class SadPecahKKViewSet(CustomView):
    queryset = SadPecahKK.objects.order_by("id").all()
    serializer_class = SadPecahKKSerializer
    permission_classes = [IsAdminUserOrReadOnly]

    def get_serializer_class(self):
        if self.action in ["update", "retrieve", "delete"]:
            raise NotFound("Operasi ini tidak tersedia")
        return SadPecahKKSerializer
