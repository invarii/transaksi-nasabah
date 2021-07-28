from io import BytesIO
import numpy as np

from django.conf import settings
from rest_framework.decorators import action
from rest_framework.response import Response
from django.http import HttpResponse
from dynamic_rest.viewsets import DynamicModelViewSet
import pytz
from openpyxl import Workbook
from rest_framework.permissions import AllowAny

from .serializers import *
from .models import *


class NasabahViewSet(DynamicModelViewSet):
    queryset = Nasabah.objects.all().order_by("id")
    serializer_class = NasabahSerializer
    permission_classes = [AllowAny]


class TransaksiViewSet(DynamicModelViewSet):
    queryset = Transaksi.objects.all().order_by("id")
    serializer_class = TransaksiSerializer
    permission_classes = [AllowAny]

    @action(detail=False, methods=["get"])
    def print(self, request):
        extras = {
            # "Nomor": "id",
            "Nasabah": "nasabah.name",
            "Deskripsi": "description",
            "Status": "debit_credit_status",
            "Tanggal": "transaction_date",
        }
        data = (
            self.get_queryset()
            .extra(select=extras, tables=("nasabah",))
            .values(*extras.keys())
            .all()
        )

        workbook = Workbook()
        sheet = workbook.active

        headers = [i for i in extras.keys()]
        for index, value in enumerate(headers):
            sheet.cell(row=1, column=index + 1).value = value

        for i, x in enumerate(data):
            for idx, value in enumerate(x.values()):
                sheet.cell(row=i + 2, column=idx + 1).value = value

        output = BytesIO()
        workbook.save(output)
        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.ms-excel",
        )
        response[
            "Content-Disposition"
        ] = 'attachment; filename="Report.xlsx"'
        return response